# ============================================
# STEP 1 - Load and Explore the Data
# ============================================
# ADD THESE 3 LINES AT THE TOP
from dotenv import load_dotenv
import os
load_dotenv()
import pandas as pd

df = pd.read_csv('tickets.csv')

print("Total tickets:", df.shape[0])
print("Total columns:", df.shape[1])
print("Column names:", df.columns.tolist())
print("\nFirst 5 rows:")
print(df.head())
print("\nMissing values:")
print(df.isnull().sum())
print("\nPriority values:")
print(df['priority'].value_counts())
print("\nQueue values:")
print(df['queue'].value_counts())

# ============================================
# STEP 2 - Clean and Prepare the Data
# ============================================
df = df.drop_duplicates()
df = df.dropna(subset=['queue', 'priority'])
df['queue'] = df['queue'].str.strip().str.title()
df['priority'] = df['priority'].str.strip().str.title()

print("\nCleaned Data Sample:")
print(df.head())
print("\nData types:")
print(df.dtypes)

# ============================================
# STEP 3 - Generate Charts
# ============================================
import matplotlib.pyplot as plt
import os

os.makedirs('charts', exist_ok=True)

# Chart 1: Tickets per Queue
queue_counts = df['queue'].value_counts()
plt.figure(figsize=(12, 6))
queue_counts.plot(kind='barh', color='steelblue')
plt.title('Tickets per Queue', fontsize=14)
plt.xlabel('Number of Tickets')
plt.tight_layout()
plt.savefig('charts/tickets_per_queue.png')
plt.close()
print("Chart 1 saved")

# Chart 2: Tickets by Priority
priority_counts = df['priority'].value_counts()
plt.figure(figsize=(7, 7))
priority_counts.plot(kind='pie', autopct='%1.1f%%', startangle=140,
                     colors=['#ff6b6b','#ffd93d','#6bcb77','#4d96ff'])
plt.title('Ticket Distribution by Priority', fontsize=14)
plt.ylabel('')
plt.tight_layout()
plt.savefig('charts/tickets_by_priority.png')
plt.close()
print("Chart 2 saved")

# Chart 3: Tickets by Language
lang_counts = df['language'].value_counts()
plt.figure(figsize=(8, 5))
lang_counts.plot(kind='bar', color='coral')
plt.title('Tickets by Language', fontsize=14)
plt.xlabel('Language')
plt.ylabel('Number of Tickets')
plt.xticks(rotation=45)
plt.tight_layout()
plt.savefig('charts/tickets_by_language.png')
plt.close()
print("Chart 3 saved")

# Chart 4: Queue vs Priority Heatmap
pivot = df.groupby(['queue', 'priority']).size().unstack(fill_value=0)
plt.figure(figsize=(12, 8))
plt.imshow(pivot.values, aspect='auto', cmap='YlOrRd')
plt.colorbar(label='Ticket Count')
plt.xticks(range(len(pivot.columns)), pivot.columns, rotation=45)
plt.yticks(range(len(pivot.index)), pivot.index)
plt.title('Queue vs Priority Heatmap', fontsize=14)
plt.tight_layout()
plt.savefig('charts/queue_priority_heatmap.png')
plt.close()
print("Chart 4 saved")

# ============================================
# STEP 4 - Generate Excel Report
# ============================================
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

wb = Workbook()

# Sheet 1: Summary
ws1 = wb.active
ws1.title = "Summary"
ws1['A1'] = "Helpdesk Ticket Analysis - Summary"
ws1['A1'].font = Font(bold=True, size=14)
ws1.merge_cells('A1:C1')
ws1['A3'] = "Metric"
ws1['B3'] = "Value"
for cell in [ws1['A3'], ws1['B3']]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="4472C4")
stats = [
    ("Total Tickets", df.shape[0]),
    ("Total Queues", df['queue'].nunique()),
    ("Total Priorities", df['priority'].nunique()),
    ("Total Languages", df['language'].nunique()),
    ("Busiest Queue", df['queue'].value_counts().idxmax()),
    ("Most Common Priority", df['priority'].value_counts().idxmax()),
]
for i, (metric, value) in enumerate(stats, start=4):
    ws1[f'A{i}'] = metric
    ws1[f'B{i}'] = value
ws1.column_dimensions['A'].width = 25
ws1.column_dimensions['B'].width = 30
print("Sheet 1 created: Summary")

# Sheet 2: Queue Breakdown
ws2 = wb.create_sheet("Queue Breakdown")
queue_df = df['queue'].value_counts().reset_index()
queue_df.columns = ['Queue', 'Ticket Count']
for r in dataframe_to_rows(queue_df, index=False, header=True):
    ws2.append(r)
for cell in ws2[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="4472C4")
ws2.column_dimensions['A'].width = 35
ws2.column_dimensions['B'].width = 15
img1 = XLImage('charts/tickets_per_queue.png')
img1.width, img1.height = 500, 300
ws2.add_image(img1, 'D2')
print("Sheet 2 created: Queue Breakdown")

# Sheet 3: Priority Breakdown
ws3 = wb.create_sheet("Priority Breakdown")
priority_df = df['priority'].value_counts().reset_index()
priority_df.columns = ['Priority', 'Ticket Count']
for r in dataframe_to_rows(priority_df, index=False, header=True):
    ws3.append(r)
for cell in ws3[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="ED7D31")
ws3.column_dimensions['A'].width = 20
ws3.column_dimensions['B'].width = 15
img2 = XLImage('charts/tickets_by_priority.png')
img2.width, img2.height = 400, 400
ws3.add_image(img2, 'D2')
print("Sheet 3 created: Priority Breakdown")

# Sheet 4: Language Breakdown
ws4 = wb.create_sheet("Language Breakdown")
lang_df = df['language'].value_counts().reset_index()
lang_df.columns = ['Language', 'Ticket Count']
for r in dataframe_to_rows(lang_df, index=False, header=True):
    ws4.append(r)
for cell in ws4[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="70AD47")
ws4.column_dimensions['A'].width = 15
ws4.column_dimensions['B'].width = 15
img3 = XLImage('charts/tickets_by_language.png')
img3.width, img3.height = 500, 300
ws4.add_image(img3, 'D2')
print("Sheet 4 created: Language Breakdown")

# Sheet 5: Raw Data
ws5 = wb.create_sheet("Raw Data")
for r in dataframe_to_rows(df, index=False, header=True):
    ws5.append(r)
for cell in ws5[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="7030A0")
for col in ws5.columns:
    ws5.column_dimensions[col[0].column_letter].width = 20
print("Sheet 5 created: Raw Data")

wb.save('helpdesk_report.xlsx')
print("✅ Excel report saved!")

# ============================================
# STEP 5 - AI-Powered Ticket Analysis (Groq)
# ============================================
from groq import Groq
import json
import time

client = Groq(api_key=os.getenv("GROQ_API_KEY"))
def analyze_ticket(subject, body):
    prompt = f"""Analyze this helpdesk ticket. Reply ONLY in JSON, no extra text.

Subject: {subject}
Body: {body}

JSON format:
{{
  "category": "Technical Issue or Billing or Compliance or Integration or Platform",
  "urgency": "Low or Medium or High or Critical",
  "suggested_action": "one short sentence"
}}"""

    response = client.chat.completions.create(
model="llama-3.1-8b-instant" ,       messages=[{"role": "user", "content": prompt}],
        max_tokens=200
    )
    raw = response.choices[0].message.content.strip()
    raw = raw.replace("```json", "").replace("```", "").strip()
    return json.loads(raw)

print("\nAnalyzing tickets with AI...\n")
ai_results = []
sample = df.head(20).copy()

for i, row in sample.iterrows():
    try:
        time.sleep(1)
        result = analyze_ticket(str(row['subject']), str(row['body']))
        result['subject'] = row['subject']
        result['queue'] = row['queue']
        result['priority'] = row['priority']
        ai_results.append(result)
        print(f"✅ Ticket {i+1}: {result['urgency']} | {result['category']}")
    except Exception as e:
        print(f"❌ Ticket {i+1} failed: {e}")

print(f"\nTotal AI results collected: {len(ai_results)}")

if len(ai_results) == 0:
    print("❌ No results. Check your API key.")
else:
    ai_df = pd.DataFrame(ai_results)
    ai_df = ai_df[['subject', 'queue', 'priority', 'category', 'urgency', 'suggested_action']]

    from openpyxl import load_workbook
    wb2 = load_workbook('helpdesk_report.xlsx')
    ws6 = wb2.create_sheet("AI Analysis")

    for r in dataframe_to_rows(ai_df, index=False, header=True):
        ws6.append(r)

    for cell in ws6[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="C00000")

    for col in ws6.columns:
        ws6.column_dimensions[col[0].column_letter].width = 25

    wb2.save('helpdesk_report.xlsx')
    print("\n✅ AI Analysis sheet added to helpdesk_report.xlsx")

    # ── Sheet 6: Manager's Executive Summary ──────────────────
ws6 = wb.create_sheet("Executive Summary")

# Filter for Critical tickets
critical_tickets = df[df['priority'].str.contains('Critical', case=False, na=False)]

if not critical_tickets.empty:
    # Prepare a list of the critical issues for the AI to summarize
    issues_list = "\n".join(critical_tickets['description'].tolist())
    
    summary_prompt = f"Summarize the following critical helpdesk issues into a single, concise paragraph for a manager. Focus on the main technical themes and urgency:\n\n{issues_list}"
    
    try:
        summary_response = client.chat.completions.create(
            model="llama-3.1-8b-instant",
            messages=[{"role": "user", "content": summary_prompt}]
        )
        summary_text = summary_response.choices[0].message.content
    except Exception as e:
        summary_text = f"Could not generate summary: {e}"

    # Formatting the Sheet
    ws6['A1'] = "Critical Issues: Executive Summary"
    ws6['A1'].font = Font(bold=True, size=14, color="FF0000")
    
    ws6['A3'] = summary_text
    ws6['A3'].alignment = Alignment(wrap_text=True, vertical='top')
    ws6.merge_cells('A3:F10') # Give the summary plenty of space
    
    print("Sheet 6 created: Executive Summary")
else:
    ws6['A1'] = "No critical tickets identified today."
    import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

def send_report_email(file_path, recipient_email):
    # Setup - Replace these with your details
    sender_email = os.getenv("SENDER_EMAIL")
    sender_password = os.getenv("GMAIL_PASSWORD")
    
    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = recipient_email
    msg['Subject'] = "Weekly Helpdesk Analysis Report"

    body = "Please find the latest AI-generated Helpdesk Analysis attached."
    msg.attach(MIMEText(body, 'plain'))

    # Attach the Excel file
    with open(file_path, "rb") as attachment:
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(attachment.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f"attachment; filename= {file_path}")
        msg.attach(part)

    # Send the email
    try:
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(sender_email, sender_password)
        server.send_message(msg)
        server.quit()
        print("✅ Email sent successfully!")
    except Exception as e:
        print(f"❌ Failed to send email: {e}")

# Call it after wb.save()
# send_report_email('helpdesk_report.xlsx', 'boss@company.com')
# This is the line that actually sends the email!
send_report_email('helpdesk_report.xlsx', 'boss@company.com') # Change to your boss's email